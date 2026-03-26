# =====================================================================
# api.py — REST API Service
# =====================================================================

from __future__ import annotations

import asyncio
import json
import hashlib
import hmac
import io
import logging
import os
import uuid
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Optional, List, Set, Dict, Any
from fastapi import FastAPI, HTTPException, Query, Request, Response, Depends, UploadFile, File, Form
from fastapi.responses import RedirectResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel

from app.auth_store import AuthStore
from app.redis_config import build_redis_client

# =====================================================================
# LOGGNING
# =====================================================================
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("api")

# =====================================================================
# KONFIGURATION
# =====================================================================
APP_SECRET_RAW = os.getenv("APP_SECRET", "").strip()
if not APP_SECRET_RAW:
    raise RuntimeError("APP_SECRET must be set")

APP_SECRET = APP_SECRET_RAW.encode("utf-8")
SESSION_TTL_MIN = int(os.getenv("SESSION_TTL_MIN", "720"))
API_PORT = int(os.getenv("API_PORT", "8000"))
SESSION_COOKIE_SECURE = os.getenv("SESSION_COOKIE_SECURE", "true").lower() in ("1", "true", "yes")
MAX_IMPORT_FILE_BYTES = int(os.getenv("MAX_IMPORT_FILE_BYTES", "2097152"))

# File paths
BASE = Path("/data")
TRANSACTIONS_FILE = BASE / "transactions.json"
AUTH_FILE = BASE / "config" / "auth_tags.json"
USERS_FILE = BASE / "config" / "users.json"
ORGS_FILE = BASE / "config" / "orgs.json"
CPS_FILE = BASE / "config" / "cps.json"
RFIDS_FILE = BASE / "config" / "rfids.json"
RFID_AUDIT_FILE = BASE / "rfid_audit.json"

# =====================================================================
# REDIS CLIENT
# =====================================================================
redis_client = build_redis_client()
auth_store = AuthStore(AUTH_FILE)


async def wait_for_redis(retries: int = 15, delay_seconds: float = 2.0):
    last_error = None
    for attempt in range(1, retries + 1):
        try:
            redis_client.ping()
            logger.info("Redis connection ready")
            return
        except Exception as exc:
            last_error = exc
            logger.warning(
                "Redis not ready yet (attempt %d/%d): %s",
                attempt,
                retries,
                exc,
            )
            await asyncio.sleep(delay_seconds)

    raise RuntimeError(f"Redis was not ready after {retries} attempts") from last_error

# =====================================================================
# HJÄLPFUNKTIONER
# =====================================================================
def load_json(path: Path, default):
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return default

def save_json(path: Path, data):
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")

def load_transactions() -> List[dict]:
    return load_json(TRANSACTIONS_FILE, [])

def save_transactions(txs: List[dict]):
    save_json(TRANSACTIONS_FILE, txs)

def load_users_map() -> dict:
    return load_json(USERS_FILE, {})

def save_users_map(d: dict):
    save_json(USERS_FILE, d)

def load_orgs() -> dict:
    return load_json(ORGS_FILE, {})

def save_orgs(d: dict):
    save_json(ORGS_FILE, d)

def load_cps_map() -> dict:
    return load_json(CPS_FILE, {})

def save_cps_map(d: dict):
    save_json(CPS_FILE, d)

def load_rfids_map() -> dict:
    return load_json(RFIDS_FILE, {})

def save_rfids_map(d: dict):
    save_json(RFIDS_FILE, d)

def load_rfid_audit() -> List[dict]:
    return load_json(RFID_AUDIT_FILE, [])

def save_rfid_audit(rows: List[dict]):
    save_json(RFID_AUDIT_FILE, rows)

def normalize_tag(tag: str) -> str:
    return (tag or "").strip().upper()

def find_user_by_email(users: dict, email: str) -> tuple[Optional[str], Optional[dict]]:
    wanted = (email or "").strip().lower()
    if not wanted:
        return None, None
    for tag, u in users.items():
        if (u.get("email") or "").strip().lower() == wanted:
            return tag, u
    return None, None

def migrate_rfids_from_users_if_needed() -> int:
    rfids = load_rfids_map()
    users = load_users_map()
    changed = 0

    for tag, user in users.items():
        ntag = normalize_tag(tag)
        if not ntag:
            continue
        if ntag not in rfids:
            rfids[ntag] = {
                "alias": ntag,
                "org_id": user.get("org_id") or "default",
                "user_email": (user.get("email") or "").strip().lower() or None,
                "active": True,
                "updated_at": iso_now(),
            }
            changed += 1

    if changed:
        save_rfids_map(rfids)
    return changed

def append_rfid_audit(actor_email: str, action: str, tag: str, details: dict):
    rows = load_rfid_audit()
    rows.append(
        {
            "at": iso_now(),
            "actor_email": (actor_email or "unknown").strip().lower(),
            "action": action,
            "tag": normalize_tag(tag),
            "details": details,
        }
    )
    # Keep file bounded to latest 5000 events.
    if len(rows) > 5000:
        rows = rows[-5000:]
    save_rfid_audit(rows)

def sync_users_for_rfid(users: dict, tag: str, user_email: Optional[str], org_id: str) -> bool:
    """Mirror RFID assignment changes into users map (keyed by RFID tag)."""
    ntag = normalize_tag(tag)
    email = (user_email or "").strip().lower() or None
    changed = False

    if not email:
        if ntag in users:
            users.pop(ntag, None)
            changed = True
        return changed

    src_tag, src_user = find_user_by_email(users, email)
    if not src_user:
        return changed

    src_tag = normalize_tag(src_tag or "") if src_tag else None
    moved = dict(src_user)
    moved["email"] = email
    moved["org_id"] = org_id or moved.get("org_id") or "default"

    if users.get(ntag) != moved:
        users[ntag] = moved
        changed = True

    if src_tag and src_tag != ntag:
        users.pop(src_tag, None)
        changed = True

    return changed

def ensure_default_org():
    """Se till att org 'default' alltid finns."""
    orgs = load_orgs()
    if "default" not in orgs:
        orgs["default"] = {"name": "Default"}
        save_orgs(orgs)

def org_for_cp(cp_id: str) -> str:
    """Returnera CP-org (om saknas → 'default')."""
    cps = load_cps_map()
    entry = cps.get(cp_id)
    if isinstance(entry, dict):
        return (entry.get("org_id") or "default").strip() or "default"
    return (entry or "default").strip() if isinstance(entry, str) else "default"

def normalize_cps_map(raw: dict) -> dict:
    """Normalize cps.json to cp_id -> {org_id, alias} while supporting legacy cp_id -> org_id format."""
    out = {}
    for cp_id, entry in (raw or {}).items():
        cpid = str(cp_id or "").strip()
        if not cpid:
            continue
        if isinstance(entry, dict):
            org_id = str(entry.get("org_id") or "default").strip() or "default"
            alias = str(entry.get("alias") or cpid).strip() or cpid
        else:
            org_id = str(entry or "default").strip() or "default"
            alias = cpid
        out[cpid] = {"org_id": org_id, "alias": alias}
    return out

def display_name_for_tag(tag: str, users_map: dict) -> str:
    u = users_map.get(tag, {})
    if not u:
        rfid = load_rfids_map().get(normalize_tag(tag), {})
        email = (rfid.get("user_email") or "").strip().lower()
        _, u = find_user_by_email(users_map, email)
        if not u:
            return rfid.get("alias") or tag
    if u.get("name"):
        return u["name"]
    fn = (u.get("first_name") or "").strip()
    ln = (u.get("last_name") or "").strip()
    return (fn + " " + ln).strip() or tag

def iso_now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.%fZ")

def utcnow() -> datetime:
    return datetime.now(timezone.utc)


def iter_redis_keys(pattern: str, *, count: int = 200):
    cursor = 0
    while True:
        cursor, keys = redis_client.scan(cursor=cursor, match=pattern, count=count)
        for key in keys:
            yield key
        if cursor == 0:
            break

# =====================================================================
# LÖSENORD / SESSIONS
# =====================================================================
def _b64(x: bytes) -> str:
    import base64
    return base64.urlsafe_b64encode(x).decode("ascii").rstrip("=")

def _b64d(s: str) -> bytes:
    import base64
    return base64.urlsafe_b64decode(s + "=" * (-len(s) % 4))

def hash_password(password: str, salt_b64: str | None = None) -> dict:
    if not salt_b64:
        salt_b64 = _b64(os.urandom(16))
    dk = hashlib.pbkdf2_hmac("sha256", password.encode(), _b64d(salt_b64), 200_000, 32)
    return {"pwd_salt": salt_b64, "pwd_hash": _b64(dk)}

def verify_password(password: str, salt_b64: str, pwh_b64: str) -> bool:
    dk = hashlib.pbkdf2_hmac("sha256", password.encode(), _b64d(salt_b64), 200_000, 32)
    return hmac.compare_digest(_b64(dk), pwh_b64)

def set_session_cookie(resp: Response, *, email: str, role: str, org_id: Optional[str]):
    token_raw = {
        "email": email,
        "role": role,
        "org_id": org_id,
        "exp": (utcnow() + timedelta(minutes=SESSION_TTL_MIN)).isoformat()
    }
    raw = json.dumps(token_raw).encode()
    sig = hmac.new(APP_SECRET, raw, hashlib.sha256).digest()
    token = f"{_b64(raw)}.{_b64(sig)}"

    resp.set_cookie(
        "session",
        token,
        httponly=True,
        samesite="Lax",
        secure=SESSION_COOKIE_SECURE,
        path="/",
        max_age=SESSION_TTL_MIN * 60,
    )

def clear_session_cookie(resp: Response):
    resp.delete_cookie("session", path="/")

def verify_token(token: str) -> dict:
    try:
        raw_b64, sig_b64 = token.split(".")
        raw = _b64d(raw_b64)
        sig = _b64d(sig_b64)
        good = hmac.new(APP_SECRET, raw, hashlib.sha256).digest()
        if not hmac.compare_digest(sig, good):
            raise ValueError("bad signature")

        data = json.loads(raw)
        exp = datetime.fromisoformat(data["exp"])
        if exp.tzinfo is None: exp = exp.replace(tzinfo=timezone.utc)
        if utcnow() > exp:
            raise ValueError("expired")
        return data
    except Exception as e:
        raise HTTPException(401, "Invalid/expired session") from e

# =====================================================================
# DEPENDENCIES (auth, roller)
# =====================================================================
def get_session(request: Request):
    token = request.cookies.get("session")
    if not token:
        raise HTTPException(401, "Not authenticated")
    return verify_token(token)

def require_auth(s=Depends(get_session)):
    return s

def require_portal_admin(s=Depends(get_session)):
    if (s.get("role") or "").lower() not in ("portal_admin", "admin"):
        raise HTTPException(403, "Portal admin required")
    return s

def require_org_admin_or_portal(s=Depends(get_session)):
    if (s.get("role") or "").lower() not in ("org_admin", "portal_admin", "admin"):
        raise HTTPException(403, "Admin/org_admin required")
    return s

# =====================================================================
# FASTAPI APP
# =====================================================================
app = FastAPI(title="EV CSMS API", version="1.0")

# ---------- MODELLER ----------
class LoginBody(BaseModel):
    email: str
    password: str

class UserMapBody(BaseModel):
    tag: str
    old_tag: Optional[str] = None
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    name: Optional[str] = None
    email: Optional[str] = None
    role: Optional[str] = None
    org_id: Optional[str] = None
    password: Optional[str] = None

class CpAssignBody(BaseModel):
    cp_id: str
    org_id: str
    alias: Optional[str] = None

class OcppCommandBody(BaseModel):
    cp_id: str
    command: str
    payload: Optional[Dict[str, Any]] = None


def _as_int(value: Any, field_name: str, *, minimum: Optional[int] = None) -> int:
    try:
        result = int(value)
    except Exception as exc:
        raise HTTPException(400, f"{field_name} måste vara ett heltal") from exc
    if minimum is not None and result < minimum:
        raise HTTPException(400, f"{field_name} måste vara minst {minimum}")
    return result


def resolve_latest_transaction_id_for_cp(cp_id: str, connector_id: int) -> int:
    cp_id = (cp_id or "").strip()
    if not cp_id:
        raise HTTPException(400, "cp_id krävs för att hitta senaste transaktionen")
    if connector_id < 1:
        raise HTTPException(400, "connector_id måste vara minst 1 för RemoteStopTransaction")

    candidates: List[dict] = []

    for key in iter_redis_keys("open_tx:*"):
        raw = redis_client.get(key)
        if not raw:
            continue
        try:
            entry = json.loads(raw.decode())
        except Exception:
            continue
        if (entry.get("charge_point") or "") != cp_id:
            continue
        if int(entry.get("connectorId") or 0) != connector_id:
            continue
        if entry.get("stop_time") or entry.get("meter_stop") is not None:
            continue
        candidates.append(entry)

    if not candidates:
        for entry in load_transactions():
            if (entry.get("charge_point") or "") != cp_id:
                continue
            if int(entry.get("connectorId") or 0) != connector_id:
                continue
            if entry.get("stop_time") or entry.get("meter_stop") is not None:
                continue
            candidates.append(entry)

    if not candidates:
        raise HTTPException(409, f"Ingen aktiv transaktion hittades för {cp_id} på uttag {connector_id}")

    latest = max(candidates, key=lambda tx: int(tx.get("transaction_id") or 0))
    tx_id = int(latest.get("transaction_id") or 0)
    if tx_id < 1:
        raise HTTPException(409, "Senaste aktiva transaktionen saknar giltigt transaction_id")
    return tx_id


def validate_ocpp_command_payload(command: str, payload: Optional[Dict[str, Any]], *, cp_id: Optional[str] = None) -> Dict[str, Any]:
    command = (command or "").strip().lower()
    payload = dict(payload or {})

    if command == "reset":
        reset_type = str(payload.get("type", "Hard")).strip() or "Hard"
        if reset_type not in {"Hard", "Soft"}:
            raise HTTPException(400, "Reset type måste vara Hard eller Soft")
        return {"type": reset_type}

    if command == "change_availability":
        availability_type = str(payload.get("type", "Operative")).strip() or "Operative"
        if availability_type not in {"Operative", "Inoperative"}:
            raise HTTPException(400, "type måste vara Operative eller Inoperative")
        return {
            "type": availability_type,
            "connector_id": _as_int(payload.get("connector_id", 0), "connector_id", minimum=0),
        }

    if command == "trigger_message":
        requested_message = str(payload.get("requested_message", "StatusNotification")).strip() or "StatusNotification"
        allowed_messages = {
            "BootNotification",
            "DiagnosticsStatusNotification",
            "FirmwareStatusNotification",
            "Heartbeat",
            "MeterValues",
            "StatusNotification",
        }
        if requested_message not in allowed_messages:
            raise HTTPException(400, f"requested_message måste vara en av: {', '.join(sorted(allowed_messages))}")
        normalized = {"requested_message": requested_message}
        if payload.get("connector_id") not in (None, ""):
            normalized["connector_id"] = _as_int(payload.get("connector_id"), "connector_id", minimum=0)
        return normalized

    if command == "clear_cache":
        return {}

    if command == "unlock_connector":
        return {"connector_id": _as_int(payload.get("connector_id", 1), "connector_id", minimum=1)}

    if command == "remote_start_transaction":
        id_tag = normalize_tag(payload.get("id_tag") or "")
        if not id_tag:
            raise HTTPException(400, "id_tag krävs för RemoteStartTransaction")
        normalized = {"id_tag": id_tag}
        if payload.get("connector_id") not in (None, ""):
            normalized["connector_id"] = _as_int(payload.get("connector_id"), "connector_id", minimum=1)
        return normalized

    if command == "remote_stop_transaction":
        connector_id = _as_int(payload.get("connector_id", 1), "connector_id", minimum=1)
        if payload.get("transaction_id") not in (None, ""):
            return {
                "connector_id": connector_id,
                "transaction_id": _as_int(payload.get("transaction_id"), "transaction_id", minimum=1),
            }
        return {
            "connector_id": connector_id,
            "transaction_id": resolve_latest_transaction_id_for_cp(cp_id or "", connector_id),
        }

    if command == "get_configuration":
        keys_raw = payload.get("key")
        if keys_raw in (None, ""):
            return {}
        if isinstance(keys_raw, str):
            keys = [k.strip() for k in keys_raw.split(",") if k.strip()]
        elif isinstance(keys_raw, list):
            keys = [str(k).strip() for k in keys_raw if str(k).strip()]
        else:
            raise HTTPException(400, "key måste vara text eller lista")
        return {"key": keys}

    raise HTTPException(400, f"Ogiltigt command: {command}")


def make_json_safe(value: Any):
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, dict):
        return {str(k): make_json_safe(v) for k, v in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [make_json_safe(v) for v in value]
    if hasattr(value, "model_dump"):
        return make_json_safe(value.model_dump())
    if hasattr(value, "dict"):
        return make_json_safe(value.dict())
    if hasattr(value, "value") and not callable(getattr(value, "value")):
        return make_json_safe(value.value)
    if hasattr(value, "__dict__"):
        return make_json_safe({k: v for k, v in vars(value).items() if not k.startswith("_")})
    return str(value)

class RfidBody(BaseModel):
    tag: str
    alias: Optional[str] = None
    org_id: Optional[str] = None
    user_email: Optional[str] = None
    active: Optional[bool] = True

class RfidPatchBody(BaseModel):
    alias: Optional[str] = None
    org_id: Optional[str] = None
    user_email: Optional[str] = None
    active: Optional[bool] = None


def normalize_import_role(value: str) -> str:
    role = (value or "user").strip().lower()
    aliases = {
        "admin": "portal_admin",
        "portal-admin": "portal_admin",
        "portal admin": "portal_admin",
        "org-admin": "org_admin",
        "org admin": "org_admin",
    }
    role = aliases.get(role, role)
    return role


def next_unset_tag(users: dict) -> str:
    tag = f"UNSET-{uuid.uuid4().hex[:8].upper()}"
    while tag in users:
        tag = f"UNSET-{uuid.uuid4().hex[:8].upper()}"
    return tag


def parse_boolish(value: Any, default: bool = False) -> bool:
    raw = str(value or "").strip().lower()
    if raw == "":
        return default
    return raw in ("1", "true", "yes", "on")


def process_import_row(
    row: dict,
    *,
    line: int,
    session: dict,
    users: dict,
    rfids: dict,
    orgs: dict,
) -> dict:
    role_session = (session.get("role") or "").lower()
    raw_tag = normalize_tag(row.get("tag") or "")
    first_name = (row.get("first_name") or "").strip()
    last_name = (row.get("last_name") or "").strip()
    name = (row.get("name") or "").strip() or f"{first_name} {last_name}".strip()
    email = (row.get("email") or "").strip().lower()
    role = normalize_import_role(row.get("role") or "user")
    password = (row.get("password") or "").strip()

    if not first_name and not last_name and not name and not email and not raw_tag:
        return {"line": line, "status": "skipped", "message": "Tom rad"}

    if not email:
        raise ValueError("email krävs")
    if "@" not in email:
        raise ValueError("email är ogiltig")
    if not name:
        raise ValueError("name eller first_name/last_name krävs")
    if role not in ("user", "org_admin", "portal_admin"):
        raise ValueError("role måste vara user, org_admin eller portal_admin")

    if role_session in ("org_admin",):
        if role == "portal_admin":
            raise ValueError("org_admin får inte skapa portal_admin")
        requested_org = (row.get("org_id") or "").strip()
        session_org = (session.get("org_id") or "").strip()
        if requested_org and requested_org != session_org:
            raise ValueError("org_admin får bara skapa användare i sin organisation")
        org_id = session_org
    else:
        org_id = (row.get("org_id") or "").strip()
        if not org_id:
            raise ValueError("org_id krävs")

    if org_id not in orgs:
        raise ValueError("okänd organisation")

    tag = raw_tag or next_unset_tag(users)
    no_rfid = not raw_tag

    for t, u in users.items():
        if t == tag:
            continue
        if (u.get("email") or "").strip().lower() == email:
            raise ValueError("email används redan")

    existing_tag_user = users.get(tag)
    if existing_tag_user:
        existing_email = (existing_tag_user.get("email") or "").strip().lower()
        if existing_email and existing_email != email:
            raise ValueError(f"taggen {tag} används redan av annan användare")

    entry = dict(existing_tag_user or {})
    entry.update(
        {
            "first_name": first_name or entry.get("first_name"),
            "last_name": last_name or entry.get("last_name"),
            "name": name,
            "email": email,
            "role": role,
            "org_id": org_id,
        }
    )
    if password:
        h = hash_password(password)
        entry["pwd_salt"] = h["pwd_salt"]
        entry["pwd_hash"] = h["pwd_hash"]

    if not no_rfid:
        r = dict(rfids.get(tag, {}))
        existing_assigned_email = (r.get("user_email") or "").strip().lower()
        if existing_assigned_email and existing_assigned_email != email:
            raise ValueError(f"RFID {tag} är redan tilldelad annan användare")

        existing_org = (r.get("org_id") or "").strip()
        if existing_org and existing_org != org_id:
            raise ValueError(f"RFID {tag} tillhör annan organisation")

        r["alias"] = (row.get("alias") or "").strip() or r.get("alias") or tag
        r["org_id"] = org_id
        r["user_email"] = email
        r["active"] = parse_boolish(row.get("active"), bool(r.get("active", True)))
        r["updated_at"] = iso_now()
        staged_rfids = dict(rfids)
        staged_rfids[tag] = r

        for other_tag, other in staged_rfids.items():
            if normalize_tag(other_tag) == normalize_tag(tag):
                continue
            if (other.get("user_email") or "").strip().lower() == email:
                other["user_email"] = None
                other["updated_at"] = iso_now()

        rfids.clear()
        rfids.update(staged_rfids)

    users[tag] = entry

    return {
        "line": line,
        "status": "ok",
        "email": email,
        "tag": tag,
        "org_id": org_id,
        "role": role,
        "created_without_rfid": no_rfid,
    }

# =====================================================================
# AUTH API
# =====================================================================
@app.post("/api/auth/login")
async def api_login(body: LoginBody, response: Response):
    users = load_users_map()

    matches = [
        (tag, u) for tag, u in users.items()
        if (u.get("email") or "").strip().lower() == body.email.strip().lower()
    ]
    if not matches:
        raise HTTPException(401, "Felaktig e‑post/lösenord")
    if len(matches) > 1:
        raise HTTPException(400, "E‑post används av flera RFID-taggar")

    tag, data = matches[0]
    salt, pwh = data.get("pwd_salt"), data.get("pwd_hash")
    if not salt or not pwh:
        raise HTTPException(401, "Lösenord ej satt")

    if not verify_password(body.password, salt, pwh):
        raise HTTPException(401, "Felaktig e‑post/lösenord")

    role = (data.get("role") or "user").lower()
    org_id = data.get("org_id") if role not in ("portal_admin", "admin") else None

    set_session_cookie(response, email=body.email.strip().lower(), role=role, org_id=org_id)
    return {"ok": True, "email": body.email}

@app.post("/api/auth/logout")
async def api_logout(resp: Response):
    clear_session_cookie(resp)
    return {"ok": True}

@app.get("/api/auth/me")
async def api_me(session=Depends(require_auth)):
    email = session["email"]
    users = load_users_map()
    name = None
    org_id = session.get("org_id")

    if not org_id:
        for _, u in users.items():
            if (u.get("email") or "").lower() == email:
                name = u.get("name") or (f"{u.get('first_name','')} {u.get('last_name','')}".strip())
                org_id = u.get("org_id")
                break
    else:
        for _, u in users.items():
            if (u.get("email") or "").lower() == email:
                name = u.get("name") or (f"{u.get('first_name','')} {u.get('last_name','')}".strip())
                break

    orgs = load_orgs()
    return {
        "email": email,
        "role": session["role"],
        "org_id": org_id,
        "org_name": orgs.get(org_id, {}).get("name"),
        "name": name
    }

# =====================================================================
# ORG API
# =====================================================================
@app.get("/api/orgs")
async def api_orgs(session=Depends(require_auth)):
    orgs = load_orgs()
    role = session.get("role")
    if role in ("portal_admin", "admin"):
        return orgs
    oid = session.get("org_id")
    return {oid: orgs.get(oid)} if oid in orgs else {}

@app.post("/api/orgs")
async def api_orgs_create(payload: dict, session=Depends(require_portal_admin)):
    org_id = (payload.get("org_id") or "").strip()
    name = (payload.get("name") or "").strip()
    if not org_id or not name:
        raise HTTPException(400, "org_id och name måste anges")
    orgs = load_orgs()
    if org_id in orgs:
        raise HTTPException(409, "Organisation finns redan")
    orgs[org_id] = {"name": name}
    save_orgs(orgs)
    return orgs

@app.patch("/api/orgs/{org_id}")
async def api_orgs_rename(org_id: str, payload: dict, session=Depends(require_portal_admin)):
    name = (payload.get("name") or "").strip()
    if not name:
        raise HTTPException(400, "name krävs")
    orgs = load_orgs()
    if org_id not in orgs:
        raise HTTPException(404, "Organisation saknas")
    orgs[org_id]["name"] = name
    save_orgs(orgs)
    return orgs

@app.delete("/api/orgs/{org_id}")
async def api_orgs_delete(org_id: str, force: bool = False, session=Depends(require_portal_admin)):
    orgs = load_orgs()
    if org_id not in orgs:
        raise HTTPException(404, "Organisation saknas")

    users = load_users_map()
    used = [t for t, u in users.items() if u.get("org_id") == org_id]
    if used and not force:
        raise HTTPException(409, f"Organisationen har {len(used)} användare")

    if used and force:
        for t in used:
            users[t]["org_id"] = "default"
        save_users_map(users)

    orgs.pop(org_id, None)
    save_orgs(orgs)
    return orgs

# =====================================================================
# CPS (CHARGE POINTS) API
# =====================================================================
@app.get("/api/cps/map")
async def api_cps_map(session=Depends(require_portal_admin)):
    return normalize_cps_map(load_cps_map())

@app.post("/api/cps/map")
async def api_cps_assign(body: CpAssignBody, session=Depends(require_portal_admin)):
    ensure_default_org()
    cp_id = body.cp_id.strip()
    org_id = body.org_id.strip()
    alias = (body.alias or "").strip()
    orgs = load_orgs()
    if org_id not in orgs:
        raise HTTPException(400, "Okänd org")
    cps = normalize_cps_map(load_cps_map())
    existing = cps.get(cp_id, {})
    cps[cp_id] = {
        "org_id": org_id,
        "alias": alias or existing.get("alias") or cp_id,
    }
    save_cps_map(cps)
    return cps

@app.delete("/api/cps/map")
async def api_cps_unassign(cp_id: str = Query(...), session=Depends(require_portal_admin)):
    cps = normalize_cps_map(load_cps_map())
    if cp_id in cps:
        cps.pop(cp_id)
        save_cps_map(cps)
    return cps

def allowed_cps_for_session(session: dict):
    """Portal_admin → alla CP. Org_admin → CP i egna orgen."""
    role = session.get("role")
    if role in ("portal_admin", "admin"):
        return None
    oid = session.get("org_id")
    cps = normalize_cps_map(load_cps_map())
    return {cp for cp, meta in cps.items() if (meta.get("org_id") or "default") == oid}

def fetch_status_map_for_cps(cps: List[str]) -> dict:
    """Return connector status map for provided CP ids."""
    status_data = {cp_id: {} for cp_id in cps}
    wanted = set(cps)
    status_keys = iter_redis_keys("connector_status:*")
    for key in status_keys:
        key_str = key.decode()
        parts = key_str.split(":")
        if len(parts) < 3:
            continue
        cp_id = parts[1]
        if cp_id not in wanted:
            continue
        connector_id = int(parts[2])
        raw = redis_client.get(key_str)
        if not raw:
            continue
        status_data.setdefault(cp_id, {})[connector_id] = json.loads(raw.decode())
    return status_data

@app.get("/api/cps")
async def api_cps(session=Depends(require_auth)):
    # Get connected CPs from Redis
    connected_cps = redis_client.smembers("connected_cps")
    ws_list = [cp.decode() for cp in connected_cps]

    # Get CPs with status from Redis
    status_keys = iter_redis_keys("connector_status:*")
    st_list = []
    for key in status_keys:
        cp_id = key.decode().split(":")[1]
        if cp_id not in st_list:
            st_list.append(cp_id)

    all_cps = sorted(set(ws_list) | set(st_list))
    cps_meta = normalize_cps_map(load_cps_map())
    allowed = allowed_cps_for_session(session)
    if allowed is None:
        visible = all_cps
    else:
        visible = [cp for cp in all_cps if cp in allowed]

    aliases = {cp: (cps_meta.get(cp, {}).get("alias") or cp) for cp in visible}
    return {"connected": visible, "aliases": aliases}

@app.get("/api/status")
async def api_status(session=Depends(require_auth)):
    allowed = allowed_cps_for_session(session)
    status_data = {}

    # Get all connector status from Redis
    status_keys = iter_redis_keys("connector_status:*")
    for key in status_keys:
        key_str = key.decode()
        parts = key_str.split(":")
        if len(parts) >= 3:
            cp_id = parts[1]
            connector_id = int(parts[2])

            if allowed is None or cp_id in allowed:
                if cp_id not in status_data:
                    status_data[cp_id] = {}
                status_data[cp_id][connector_id] = json.loads(redis_client.get(key_str).decode())

    return status_data

@app.get("/api/portal/live/chargers")
async def api_portal_live_chargers(org_id: Optional[str] = None, session=Depends(require_portal_admin)):
    cps_map = normalize_cps_map(load_cps_map())
    connected_cps = redis_client.smembers("connected_cps")
    connected = sorted(cp.decode() for cp in connected_cps)

    if org_id:
        connected = [cp for cp in connected if (cps_map.get(cp, {}).get("org_id") or "default") == org_id]

    status_data = fetch_status_map_for_cps(connected)

    items = []
    for cp_id in connected:
        items.append(
            {
                "cp_id": cp_id,
                "alias": cps_map.get(cp_id, {}).get("alias") or cp_id,
                "org_id": cps_map.get(cp_id, {}).get("org_id") or "default",
                "status": status_data.get(cp_id, {}),
            }
        )

    return {
        "generated_at": iso_now(),
        "items": items,
    }

@app.post("/api/portal/ocpp/command")
async def api_portal_ocpp_command(body: OcppCommandBody, session=Depends(require_portal_admin)):
    cp_id = (body.cp_id or "").strip()
    command = (body.command or "").strip().lower()

    if not cp_id:
        raise HTTPException(400, "cp_id krävs")
    if not command:
        raise HTTPException(400, "command krävs")

    allowed_commands = {
        "reset",
        "change_availability",
        "trigger_message",
        "clear_cache",
        "unlock_connector",
        "remote_start_transaction",
        "remote_stop_transaction",
        "get_configuration",
    }
    if command not in allowed_commands:
        raise HTTPException(400, f"Ogiltigt command. Tillåtna: {', '.join(sorted(allowed_commands))}")

    payload = validate_ocpp_command_payload(command, body.payload, cp_id=cp_id)

    connected = {cp.decode() for cp in redis_client.smembers("connected_cps")}
    if cp_id not in connected:
        raise HTTPException(409, "Laddare är inte ansluten")

    command_id = str(uuid.uuid4())
    user_email = session.get("email", "unknown")

    envelope = {
        "command_id": command_id,
        "cp_id": cp_id,
        "command": command,
        "payload": payload,
        "requested_by": user_email,
        "requested_at": iso_now(),
    }

    result_key = f"ocpp:command_result:{command_id}"
    redis_client.setex(
        result_key,
        600,
        json.dumps({
            "command_id": command_id,
            "status": "queued",
            "cp_id": cp_id,
            "command": command,
            "requested_at": envelope["requested_at"],
        }),
    )

    redis_client.rpush("ocpp:commands", json.dumps(envelope))
    return {"ok": True, "command_id": command_id, "status": "queued"}

@app.get("/api/portal/ocpp/command/{command_id}")
async def api_portal_ocpp_command_status(command_id: str, session=Depends(require_portal_admin)):
    raw = redis_client.get(f"ocpp:command_result:{command_id}")
    if not raw:
        raise HTTPException(404, "Kommandoresultat saknas eller har löpt ut")
    return json.loads(raw.decode())

# =====================================================================
# USERS API
# =====================================================================
@app.get("/api/users/map")
async def api_users_map(session=Depends(require_auth)):
    users = load_users_map()
    rfids = load_rfids_map()
    if rfids:
        by_rfid = {}
        for tag, item in rfids.items():
            email = (item.get("user_email") or "").strip().lower()
            if not email:
                continue
            _, user = find_user_by_email(users, email)
            if not user:
                continue
            row = dict(user)
            row["email"] = email
            row["org_id"] = item.get("org_id") or row.get("org_id") or "default"
            row["rfid_alias"] = item.get("alias") or normalize_tag(tag)
            by_rfid[normalize_tag(tag)] = row
        users = by_rfid
    role = session.get("role")

    if role in ("portal_admin", "admin"):
        return users
    elif role == "org_admin":
        oid = session.get("org_id")
        return {t: u for t, u in users.items() if u.get("org_id") == oid}
    else:
        email = (session.get("email") or "").strip().lower()
        mine = {}
        for t, u in users.items():
            if (u.get("email") or "").lower() == email:
                mine[t] = u
                break
        return mine

@app.get("/api/users/unassigned")
async def api_users_unassigned(session=Depends(require_org_admin_or_portal)):
    role = (session.get("role") or "").lower()
    scope_org = session.get("org_id")
    users = load_users_map()
    rfids = load_rfids_map()

    assigned_emails = {
        (item.get("user_email") or "").strip().lower()
        for item in rfids.values()
        if (item.get("user_email") or "").strip()
    }

    result = {}
    for tag, user in users.items():
        email = (user.get("email") or "").strip().lower()
        if email in assigned_emails:
            continue
        user_org = user.get("org_id") or "default"
        if role == "org_admin" and user_org != scope_org:
            continue
        result[normalize_tag(tag)] = user

    return result

@app.get("/api/rfids")
async def api_rfids(org_id: Optional[str] = None, assigned: Optional[bool] = None, session=Depends(require_org_admin_or_portal)):
    role = (session.get("role") or "").lower()
    scope_org = session.get("org_id")
    users = load_users_map()
    rfids = load_rfids_map()
    orgs = load_orgs()

    rows = []
    for tag, item in rfids.items():
        item_org = item.get("org_id") or "default"
        if role == "org_admin" and item_org != scope_org:
            continue
        if org_id and item_org != org_id:
            continue

        email = (item.get("user_email") or "").strip().lower() or None
        _, user = find_user_by_email(users, email or "")
        if assigned is True and not email:
            continue
        if assigned is False and email:
            continue

        rows.append(
            {
                "tag": normalize_tag(tag),
                "alias": item.get("alias") or normalize_tag(tag),
                "org_id": item_org,
                "org_name": orgs.get(item_org, {}).get("name"),
                "user_email": email,
                "user_name": (
                    user.get("name")
                    or f"{user.get('first_name', '')} {user.get('last_name', '')}".strip()
                    if user
                    else None
                ),
                "active": bool(item.get("active", True)),
                "updated_at": item.get("updated_at"),
            }
        )

    rows.sort(key=lambda r: (r["org_id"], r["alias"], r["tag"]))
    return {"items": rows, "count": len(rows)}

@app.get("/api/rfids/audit")
async def api_rfids_audit(limit: int = 200, session=Depends(require_org_admin_or_portal)):
    role = (session.get("role") or "").lower()
    scope_org = session.get("org_id")
    rows = load_rfid_audit()
    rfids = load_rfids_map()

    if role == "org_admin":
        filtered = []
        for row in rows:
            tag = normalize_tag(row.get("tag") or "")
            entry = rfids.get(tag, {})
            if (entry.get("org_id") or "default") == scope_org:
                filtered.append(row)
        rows = filtered

    limit = max(1, min(1000, int(limit)))
    return {"items": list(reversed(rows[-limit:])), "count": min(len(rows), limit)}

@app.post("/api/rfids")
async def api_rfids_create(body: RfidBody, session=Depends(require_org_admin_or_portal)):
    role = (session.get("role") or "").lower()
    actor = session.get("email") or "unknown"
    tag = normalize_tag(body.tag)
    if not tag:
        raise HTTPException(400, "tag krävs")

    rfids = load_rfids_map()
    if tag in rfids:
        raise HTTPException(409, "RFID-tagg finns redan")

    org_id = (body.org_id or "").strip() or (session.get("org_id") if role == "org_admin" else "default")
    orgs = load_orgs()
    if role == "org_admin":
        org_id = session.get("org_id")
    if org_id not in orgs:
        raise HTTPException(400, "Okänd organisation")

    users = load_users_map()
    email = (body.user_email or "").strip().lower() or None
    if email:
        _, user = find_user_by_email(users, email)
        if not user:
            raise HTTPException(404, "Användare hittades inte")
        if role == "org_admin" and user.get("org_id") != session.get("org_id"):
            raise HTTPException(403, "Kan bara tilldela användare i din organisation")
        if (user.get("org_id") or "default") != org_id:
            raise HTTPException(409, "Tagg och användare måste tillhöra samma organisation")

    rfids[tag] = {
        "alias": (body.alias or "").strip() or tag,
        "org_id": org_id,
        "user_email": email,
        "active": True if body.active is None else bool(body.active),
        "updated_at": iso_now(),
    }
    save_rfids_map(rfids)
    if sync_users_for_rfid(users, tag, email, org_id):
        save_users_map(users)
    if not auth_store.contains(tag):
        auth_store.add(tag)

    append_rfid_audit(actor, "create", tag, {"org_id": org_id, "user_email": email, "alias": rfids[tag]["alias"]})
    return {"ok": True, "tag": tag}

@app.patch("/api/rfids/{tag}")
async def api_rfids_patch(tag: str, body: RfidPatchBody, session=Depends(require_org_admin_or_portal)):
    role = (session.get("role") or "").lower()
    actor = session.get("email") or "unknown"
    tag = normalize_tag(tag)

    rfids = load_rfids_map()
    entry = rfids.get(tag)
    if not entry:
        raise HTTPException(404, "RFID-tagg saknas")

    current_org = entry.get("org_id") or "default"
    if role == "org_admin" and current_org != session.get("org_id"):
        raise HTTPException(403, "Kan bara hantera taggar i din organisation")

    users = load_users_map()
    if body.alias is not None:
        entry["alias"] = (body.alias or "").strip() or tag

    if "user_email" in body.model_fields_set:
        email = (body.user_email or "").strip().lower() or None
        if email:
            _, user = find_user_by_email(users, email)
            if not user:
                raise HTTPException(404, "Användare hittades inte")
            user_org = user.get("org_id") or "default"
            if role == "org_admin" and user_org != session.get("org_id"):
                raise HTTPException(403, "Kan bara tilldela användare i din organisation")
            if user_org != (entry.get("org_id") or "default"):
                raise HTTPException(409, "Tagg och användare måste tillhöra samma organisation")
        entry["user_email"] = email

    if body.org_id is not None:
        if role != "portal_admin" and role != "admin":
            raise HTTPException(403, "Bara portal_admin får flytta taggar mellan organisationer")
        wanted_org = (body.org_id or "").strip()
        if wanted_org not in load_orgs():
            raise HTTPException(400, "Okänd organisation")
        entry["org_id"] = wanted_org
        # Unassign if moved to a different org than the current user.
        if entry.get("user_email"):
            _, u = find_user_by_email(users, entry.get("user_email"))
            if not u or (u.get("org_id") or "default") != wanted_org:
                entry["user_email"] = None

    if body.active is not None:
        entry["active"] = bool(body.active)

    assignment_touched = ("user_email" in body.model_fields_set) or (body.org_id is not None)
    if assignment_touched and sync_users_for_rfid(users, tag, entry.get("user_email"), entry.get("org_id") or "default"):
        save_users_map(users)

    entry["updated_at"] = iso_now()
    rfids[tag] = entry
    save_rfids_map(rfids)

    if entry.get("active", True):
        if not auth_store.contains(tag):
            auth_store.add(tag)
    elif auth_store.contains(tag):
        auth_store.remove(tag)

    append_rfid_audit(actor, "update", tag, {"entry": entry})
    return {"ok": True, "tag": tag, "entry": entry}

@app.delete("/api/rfids/{tag}")
async def api_rfids_delete(tag: str, session=Depends(require_org_admin_or_portal)):
    role = (session.get("role") or "").lower()
    actor = session.get("email") or "unknown"
    tag = normalize_tag(tag)
    rfids = load_rfids_map()
    entry = rfids.get(tag)
    if not entry:
        raise HTTPException(404, "RFID-tagg saknas")

    if role == "org_admin" and (entry.get("org_id") or "default") != session.get("org_id"):
        raise HTTPException(403, "Kan bara ta bort taggar i din organisation")

    rfids.pop(tag, None)
    save_rfids_map(rfids)
    keep_in_allowlist = False
    if tag in rfids and bool(rfids[tag].get("active", True)):
        keep_in_allowlist = True

    if auth_store.contains(tag) and not keep_in_allowlist:
        auth_store.remove(tag)

    append_rfid_audit(actor, "delete", tag, {"old": entry})
    return {"ok": True}

@app.get("/api/rfids/import/template")
@app.get("/api/rfids/import/template.xlsx")
async def api_rfids_import_template(session=Depends(require_org_admin_or_portal)):
    role = (session.get("role") or "").lower()
    org_id = session.get("org_id") or "default"

    wb = Workbook()
    ws = wb.active
    ws.title = "rfid_import"

    header = ["tag", "alias", "org_id", "active"]
    ws.append(header)

    if role == "org_admin":
        ws.append(["ABC12345", "S1", org_id, "true"])
        ws.append(["DEF67890", "S2", org_id, "true"])
    else:
        ws.append(["ABC12345", "S1", "default", "true"])
        ws.append(["DEF67890", "S2", "", "true"])
        ws.append(["GHI11111", "S3", "Takorama_Storås", "true"])

    output = io.BytesIO()
    wb.save(output)

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=rfid_import_template.xlsx"},
    )


@app.post("/api/rfids/import/xlsx")
async def api_rfids_import_xlsx(
    file: UploadFile = File(...),
    dry_run: bool = Form(False),
    session=Depends(require_org_admin_or_portal),
):
    role = (session.get("role") or "").lower()
    actor = session.get("email") or "unknown"
    session_org = session.get("org_id") or "default"

    if not file.filename:
        raise HTTPException(400, "XLSX-fil saknas")

    try:
        raw = await file.read()
        if len(raw) > MAX_IMPORT_FILE_BYTES:
            raise HTTPException(413, f"Filen är för stor (max {MAX_IMPORT_FILE_BYTES} bytes)")
        wb = load_workbook(filename=io.BytesIO(raw), data_only=True)
    except Exception as exc:
        if isinstance(exc, HTTPException):
            raise
        raise HTTPException(400, "Kunde inte läsa XLSX-filen") from exc

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        raise HTTPException(400, "XLSX saknar header")

    fieldnames = [str(x).strip() if x is not None else "" for x in header_row]
    if "tag" not in fieldnames:
        raise HTTPException(400, "XLSX saknar kolumnen 'tag'")

    rfids = load_rfids_map()
    orgs  = load_orgs()

    results = []
    success = 0
    failed  = 0
    skipped = 0

    new_rfids: Dict[str, Any] = {}

    for idx, values in enumerate(rows_iter, start=2):
        row: Dict[str, str] = {
            fieldnames[i]: "" if v is None else str(v).strip()
            for i, v in enumerate(values)
            if i < len(fieldnames) and fieldnames[i]
        }

        try:
            tag = normalize_tag(row.get("tag") or "")
            if not tag:
                skipped += 1
                results.append({"line": idx, "status": "skipped", "message": "Tom tagg, hoppar över"})
                continue

            if tag in rfids or tag in new_rfids:
                skipped += 1
                results.append({"line": idx, "status": "skipped", "tag": tag, "message": f"Tagg {tag} finns redan"})
                continue

            raw_org = (row.get("org_id") or "").strip()
            if role == "org_admin":
                org_id = session_org
            else:
                org_id = raw_org if raw_org else None  # None = unassigned / undefined

            if org_id and org_id not in orgs:
                failed += 1
                results.append({"line": idx, "status": "error", "tag": tag, "message": f"Okänd organisation: {org_id}"})
                continue

            alias  = (row.get("alias") or "").strip() or tag
            active = parse_boolish(row.get("active"), True)

            entry: Dict[str, Any] = {
                "alias":      alias,
                "org_id":     org_id or "default",
                "user_email": None,
                "active":     active,
                "updated_at": iso_now(),
            }

            new_rfids[tag] = entry
            success += 1
            results.append({"line": idx, "status": "ok", "tag": tag, "message": f"Tagg {tag} importerad (org: {entry['org_id']})"})

        except Exception as exc:
            failed += 1
            results.append({"line": idx, "status": "error", "message": str(exc)})

    if not dry_run and new_rfids:
        rfids.update(new_rfids)
        save_rfids_map(rfids)
        for tag, entry in new_rfids.items():
            if entry.get("active", True) and not auth_store.contains(tag):
                auth_store.add(tag)
            append_rfid_audit(actor, "import", tag, {"org_id": entry["org_id"], "alias": entry["alias"]})

    return {
        "ok": True,
        "dry_run": bool(dry_run),
        "summary": {
            "total_rows": len(results),
            "imported":   success,
            "failed":     failed,
            "skipped":    skipped,
        },
        "results": results,
    }


@app.post("/api/users/map")
async def api_users_map_add(body: UserMapBody, req: Request, session=Depends(require_org_admin_or_portal)):
    raw = {}
    try:
        raw = await req.json()
    except:
        pass

    users = load_users_map()
    orgs = load_orgs()

    email = ((raw.get("email") or body.email or "")).strip().lower()
    if not email:
        raise HTTPException(400, "email krävs")

    full_name = body.name or " ".join([p for p in [body.first_name, body.last_name] if p]).strip()
    if not full_name:
        raise HTTPException(400, "name eller first/last_name krävs")

    role = (body.role or "user").lower()
    if role not in ("user", "org_admin", "portal_admin"):
        raise HTTPException(400, "Ogiltig roll")

    # Portal-admin kan välja org_id. Org_admin är låst till sin egen org.
    if session.get("role") in ("portal_admin", "admin"):
        org_id = (body.org_id or raw.get("org_id") or "").strip()
        if not org_id:
            raise HTTPException(400, "org_id krävs för portal_admin")
        if org_id not in orgs:
            raise HTTPException(400, "Okänd organisation")
    else:
        org_id = session.get("org_id")
        if role == "portal_admin":
            raise HTTPException(403, "org_admin får inte skapa portal_admin")

    tag = normalize_tag(body.tag or "")
    old_tag = normalize_tag(raw.get("old_tag") or body.old_tag or "")
    no_rfid = not tag
    if no_rfid:
        tag = f"UNSET-{uuid.uuid4().hex[:8].upper()}"

    # E‑post måste vara unik utanför de taggar som deltar i samma uppdatering.
    excluded_tags = {tag}
    if old_tag:
        excluded_tags.add(old_tag)
    for t, u in users.items():
        if t not in excluded_tags and (u.get("email") or "").lower() == email:
            raise HTTPException(400, "E‑post används redan")

    source_entry = users.get(tag, {})
    if old_tag and old_tag != tag:
        old_entry = users.get(old_tag)
        if not old_entry:
            # Fallback: if the UI row is built from stale RFID mapping,
            # resolve the current source user record by email.
            found_tag, found_user = find_user_by_email(users, email)
            if not found_user:
                raise HTTPException(404, "old_tag saknas")
            old_tag = normalize_tag(found_tag or "")
            old_entry = found_user
        if session.get("role") == "org_admin" and old_entry.get("org_id") != session.get("org_id"):
            raise HTTPException(403, "Får bara flytta användare i din egen organisation")
        source_entry = dict(old_entry)

    entry = dict(source_entry)
    entry.update({
        "first_name": body.first_name or entry.get("first_name"),
        "last_name": body.last_name or entry.get("last_name"),
        "name": full_name,
        "email": email,
        "role": role,
        "org_id": org_id,
    })
    if body.password:
        h = hash_password(body.password)
        entry["pwd_salt"] = h["pwd_salt"]
        entry["pwd_hash"] = h["pwd_hash"]

    users[tag] = entry
    if old_tag and old_tag != tag:
        users.pop(old_tag, None)
    save_users_map(users)

    if not no_rfid:
        rfids = load_rfids_map()
        r = rfids.get(tag, {})
        r["alias"] = r.get("alias") or tag
        r["org_id"] = org_id
        r["user_email"] = email
        r["active"] = bool(r.get("active", True))
        r["updated_at"] = iso_now()
        rfids[tag] = r

        # Ensure the same user email is assigned to only one RFID tag.
        for t, item in rfids.items():
            if normalize_tag(t) == normalize_tag(tag):
                continue
            if (item.get("user_email") or "").strip().lower() == email:
                item["user_email"] = None
                item["updated_at"] = iso_now()

        save_rfids_map(rfids)

        if not auth_store.contains(tag):
            auth_store.add(tag)

    return users

@app.delete("/api/users/map")
async def api_users_map_del(tag: str = Query(...), revoke: bool = Query(True), session=Depends(require_org_admin_or_portal)):
    tag = normalize_tag(tag)
    users = load_users_map()
    entry = users.get(tag)
    if not entry:
        return users

    if session.get("role") == "org_admin" and entry.get("org_id") != session.get("org_id"):
        raise HTTPException(403, "Får bara ta bort användare i din egen organisation")

    if revoke and entry.get("role") in ("portal_admin", "admin"):
        raise HTTPException(403, "Kan inte ta bort portal_admin")

    if revoke:
        users.pop(tag)
        save_users_map(users)
    else:
        if not tag.startswith("UNSET-"):
            unset_tag = f"UNSET-{uuid.uuid4().hex[:8].upper()}"
            while unset_tag in users:
                unset_tag = f"UNSET-{uuid.uuid4().hex[:8].upper()}"
            users[unset_tag] = entry
            users.pop(tag, None)
            save_users_map(users)

    rfids = load_rfids_map()
    if tag in rfids:
        rfids[tag]["user_email"] = None
        rfids[tag]["updated_at"] = iso_now()
        save_rfids_map(rfids)

    if revoke and auth_store.contains(tag):
        auth_store.remove(tag)

    return users


@app.get("/api/users/import/template")
@app.get("/api/users/import/template.xlsx")
async def api_users_import_template(session=Depends(require_org_admin_or_portal)):
    role = (session.get("role") or "").lower()
    org_id = session.get("org_id") or "default"
    rows = [
        ["first_name", "last_name", "email", "role", "org_id", "tag", "password", "active"],
    ]

    if role == "org_admin":
        rows.append(["Anna", "Andersson", "anna@example.com", "user", org_id, "", "", "true"])
        rows.append(["Erik", "Ek", "erik@example.com", "org_admin", org_id, "ABC12345", "", "true"])
    else:
        rows.append(["Anna", "Andersson", "anna@example.com", "user", "default", "", "", "true"])
        rows.append(["Erik", "Ek", "erik@example.com", "org_admin", "default", "ABC12345", "", "true"])

    wb = Workbook()
    ws = wb.active
    ws.title = "users_import"
    for row in rows:
        ws.append(row)

    output = io.BytesIO()
    wb.save(output)

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=users_import_template.xlsx"},
    )


@app.post("/api/users/import/xlsx")
@app.post("/api/users/import/csv")
async def api_users_import_csv(
    file: UploadFile = File(...),
    dry_run: bool = Form(False),
    session=Depends(require_org_admin_or_portal),
):
    if not file.filename:
        raise HTTPException(400, "XLSX-fil saknas")

    try:
        raw = await file.read()
        if len(raw) > MAX_IMPORT_FILE_BYTES:
            raise HTTPException(413, f"Filen är för stor (max {MAX_IMPORT_FILE_BYTES} bytes)")
        wb = load_workbook(filename=io.BytesIO(raw), data_only=True)
    except Exception as exc:
        if isinstance(exc, HTTPException):
            raise
        raise HTTPException(400, "Kunde inte läsa XLSX-filen") from exc

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        raise HTTPException(400, "XLSX saknar header")

    fieldnames = [str(x).strip() if x is not None else "" for x in header_row]
    if not any(fieldnames):
        raise HTTPException(400, "XLSX saknar header")

    required_headers = {"first_name", "last_name", "email"}
    missing = [h for h in required_headers if h not in set(fieldnames)]
    if missing:
        raise HTTPException(400, f"XLSX saknar kolumner: {', '.join(missing)}")

    users = load_users_map()
    rfids = load_rfids_map()
    orgs = load_orgs()

    results = []
    success = 0
    failed = 0
    skipped = 0

    for idx, values in enumerate(rows_iter, start=2):
        row = {
            fieldnames[col_idx]: "" if val is None else str(val)
            for col_idx, val in enumerate(values)
            if col_idx < len(fieldnames) and fieldnames[col_idx]
        }
        try:
            result = process_import_row(
                row,
                line=idx,
                session=session,
                users=users,
                rfids=rfids,
                orgs=orgs,
            )
            results.append(result)
            if result["status"] == "ok":
                success += 1
                tag = normalize_tag(result.get("tag") or "")
                active = parse_boolish(row.get("active"), True)
                if not dry_run and tag and not result.get("created_without_rfid") and active and not auth_store.contains(tag):
                    auth_store.add(tag)
            else:
                skipped += 1
        except ValueError as exc:
            failed += 1
            results.append({"line": idx, "status": "error", "message": str(exc)})

    if not dry_run and success:
        save_users_map(users)
        save_rfids_map(rfids)

    return {
        "ok": True,
        "dry_run": bool(dry_run),
        "summary": {
            "total_rows": len(results),
            "imported": success,
            "failed": failed,
            "skipped": skipped,
        },
        "results": results,
    }

# =====================================================================
# SUMMARY & HISTORY
# =====================================================================
def _allowed_tags_for_session(session: dict, users_map: dict) -> Optional[Set[str]]:
    rfids = load_rfids_map()
    role = session.get("role")
    if role in ("portal_admin", "admin"):
        return None
    if role == "org_admin":
        oid = session.get("org_id")
        return {normalize_tag(t) for t, r in rfids.items() if (r.get("org_id") or "default") == oid}
    email = (session.get("email") or "")
    return {normalize_tag(t) for t, r in rfids.items() if (r.get("user_email") or "").strip().lower() == email.strip().lower()}


def _history_rows_for_session(days: int, tag: Optional[str], session: dict) -> List[dict]:
    now = datetime.now(timezone.utc)
    cutoff = now - timedelta(days=days)
    txs = load_transactions()
    users_map = load_users_map()
    allowed = _allowed_tags_for_session(session, users_map)

    rows = []
    for tx in txs:
        stop = tx.get("stop_time")
        if not stop or tx.get("meter_stop") is None:
            continue
        stop_dt = datetime.fromisoformat(stop.replace("Z", "+00:00"))
        if stop_dt < cutoff:
            continue

        tg = normalize_tag(tx.get("id_tag") or "")
        if tag and tg != normalize_tag(tag):
            continue
        if allowed is not None and tg not in allowed:
            continue

        e = (tx["meter_stop"] - tx["meter_start"]) / 1000.0
        rows.append({
            "tag": tg,
            "name": display_name_for_tag(tg, users_map),
            "charge_point": tx.get("charge_point"),
            "connectorId": tx.get("connectorId"),
            "start_time": tx.get("start_time"),
            "stop_time": tx.get("stop_time"),
            "energy_kwh": round(max(0.0, e), 3),
        })

    rows.sort(key=lambda r: r["stop_time"] or "", reverse=True)
    return rows

@app.get("/api/users/summary")
async def api_users_summary(days: int = 30, session=Depends(require_auth)):
    now = datetime.now(timezone.utc)
    cutoff = now - timedelta(days=days)
    txs = load_transactions()
    users_map = load_users_map()
    allowed = _allowed_tags_for_session(session, users_map)

    summary = {}
    for tx in txs:
        stop = tx.get("stop_time")
        if not stop or tx.get("meter_stop") is None:
            continue
        stop_dt = datetime.fromisoformat(stop.replace("Z", "+00:00"))
        if stop_dt < cutoff:
            continue

        tag = normalize_tag(tx.get("id_tag") or "")
        if allowed is not None and tag not in allowed:
            continue

        e = (tx["meter_stop"] - tx["meter_start"]) / 1000.0
        nm = display_name_for_tag(tag, users_map)
        row = summary.setdefault(tag, {"kwh": 0.0, "sessions": 0, "name": nm})
        row["kwh"] += max(0.0, e)
        row["sessions"] += 1

    for r in summary.values():
        r["kwh"] = round(r["kwh"], 3)

    return {"period_days": days, "generated_at": iso_now(), "users": summary}

@app.get("/api/users/history")
async def api_users_history(days: int = 30, tag: Optional[str] = None, session=Depends(require_auth)):
    rows = _history_rows_for_session(days, tag, session)
    return {"period_days": days, "items": rows, "count": len(rows)}


@app.get("/api/users/history/export.xlsx")
async def api_users_history_export_xlsx(days: int = 30, tag: Optional[str] = None, session=Depends(require_auth)):
    rows = _history_rows_for_session(days, tag, session)

    per_user: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        key = (row.get("tag") or "").strip() or f"__unknown__:{(row.get('name') or '').strip()}"
        current = per_user.get(key, {
            "name": (row.get("name") or "").strip(),
            "tag": (row.get("tag") or "").strip(),
            "energy": 0.0,
            "charges": 0,
        })
        if not current["name"]:
            current["name"] = (row.get("name") or "").strip() or current["tag"] or "Unknown"
        if not current["tag"]:
            current["tag"] = (row.get("tag") or "").strip()
        current["energy"] += float(row.get("energy_kwh") or 0.0)
        current["charges"] += 1
        per_user[key] = current

    summary_rows = sorted(per_user.values(), key=lambda x: x["name"])

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append([
        "Name",
        "RFID tag",
        "Total kWh consumed during selected period",
        "Total times charged during selected period",
    ])
    for item in summary_rows:
        ws.append([
            item["name"],
            item["tag"],
            round(float(item["energy"]), 3),
            int(item["charges"]),
        ])

    output = io.BytesIO()
    wb.save(output)

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    org_suffix = (session.get("org_id") or "org").replace(" ", "_")
    filename = f"historik_{org_suffix}_{ts}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

# =====================================================================
# MY SUMMARY
# =====================================================================
@app.get("/api/my/summary")
async def api_my_summary(days: int = 30, session=Depends(require_auth)):
    now = datetime.now(timezone.utc)
    cutoff = now - timedelta(days=days)

    txs = load_transactions()
    email = session.get("email")
    rfids = load_rfids_map()
    my_tags = {normalize_tag(t) for t, r in rfids.items() if (r.get("user_email") or "").strip().lower() == (email or "").strip().lower()}

    total = 0.0
    count = 0

    for tx in txs:
        stop = tx.get("stop_time")
        if not stop or tx.get("meter_stop") is None:
            continue
        stop_dt = datetime.fromisoformat(stop.replace("Z", "+00:00"))
        if stop_dt < cutoff:
            continue

        if normalize_tag(tx.get("id_tag") or "") not in my_tags:
            continue

        total += max(0.0, (tx["meter_stop"] - tx["meter_start"]) / 1000.0)
        count += 1

    return {
        "period_days": days,
        "generated_at": iso_now(),
        "kwh": round(total, 3),
        "sessions": count
    }

# =====================================================================
# HEALTH CHECK
# =====================================================================
@app.get("/health")
async def health_check():
    return {"status": "healthy", "service": "api"}

# =====================================================================
# STARTUP
# =====================================================================
@app.on_event("startup")
async def startup():
    await wait_for_redis()
    ensure_default_org()
    RFIDS_FILE.parent.mkdir(parents=True, exist_ok=True)
    if not RFIDS_FILE.exists():
        save_rfids_map({})
    if not RFID_AUDIT_FILE.exists():
        save_rfid_audit([])

    migrated = migrate_rfids_from_users_if_needed()
    added = 0
    for tag, item in load_rfids_map().items():
        if not bool(item.get("active", True)):
            continue
        if not auth_store.contains(tag):
            auth_store.add(tag)
            added += 1
    if added:
        logger.info("Allowlist synkad: lade till %s taggar", added)
    if migrated:
        logger.info("RFID-migrering: skapade %s poster från users.json", migrated)
    logger.info("API service started on port %d", API_PORT)

# =====================================================================
# ENTRYPOINT
# =====================================================================
if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=API_PORT, log_level="info")

